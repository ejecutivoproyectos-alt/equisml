import random
import os
import re
import tempfile
import unicodedata
import calendar
import xml.etree.ElementTree as ET

from io import BytesIO
from datetime import datetime, date, timedelta

import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================================================
# CONFIGURACIÓN GENERAL
# =========================================================
st.set_page_config(page_title="Herramientas de Excel y XML", layout="wide")

THIN_BLACK = Side(style="thin", color="000000")


# =========================================================
# UTILIDADES GENERALES
# =========================================================
def normalizar_texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def centrar(cell, bold=False, size=12):
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.font = Font(bold=bold, size=size)


def aplicar_borde_rango(ws, r1, c1, r2, c2):
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for cell in row:
            cell.border = Border(
                left=THIN_BLACK,
                right=THIN_BLACK,
                top=THIN_BLACK,
                bottom=THIN_BLACK
            )


def ajustar_ancho(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def limpiar_hoja_si_existe(wb, nombre_hoja):
    if nombre_hoja in wb.sheetnames:
        del wb[nombre_hoja]


def safe_sheet_name(name):
    invalid = ['\\', '/', '*', '[', ']', ':', '?']
    for ch in invalid:
        name = name.replace(ch, "")
    return name[:31]


# =========================================================
# MÓDULO 1: XML CFDI A EXCEL
# =========================================================
def limpiar_concepto(texto):
    if not texto:
        return ""
    texto = texto.strip()
    if texto.startswith("-"):
        texto = texto.lstrip("-").strip()
    return texto


def formatear_fecha(fecha_raw):
    if not fecha_raw:
        return ""
    try:
        fecha_dt = datetime.fromisoformat(fecha_raw)
        return f"{fecha_dt.day}/{fecha_dt.month}/{fecha_dt.year}"
    except Exception:
        try:
            solo_fecha = fecha_raw.split("T")[0]
            anio, mes, dia = solo_fecha.split("-")
            return f"{int(dia)}/{int(mes)}/{anio}"
        except Exception:
            return fecha_raw


def parse_xml(file):
    registros = []

    try:
        tree = ET.parse(file)
        root = tree.getroot()

        ns = {
            "cfdi": "http://www.sat.gob.mx/cfd/4",
            "tfd": "http://www.sat.gob.mx/TimbreFiscalDigital"
        }

        fecha_raw = root.attrib.get("Fecha", "")
        total = root.attrib.get("Total", "0")
        fecha = formatear_fecha(fecha_raw)

        try:
            importe_total = float(total)
        except Exception:
            importe_total = 0.0

        timbre = root.find(".//tfd:TimbreFiscalDigital", ns)
        folio_uuid = ""
        if timbre is not None:
            folio_uuid = timbre.attrib.get("UUID", "")

        conceptos = root.findall(".//cfdi:Concepto", ns)

        for concepto in conceptos:
            descripcion = limpiar_concepto(concepto.attrib.get("Descripcion", ""))

            registros.append({
                "CONCEPTO": descripcion,
                "FECHA": fecha,
                "FOLIO": folio_uuid,
                "IMPORTE": importe_total
            })

    except Exception as e:
        st.error(f"Error al leer el archivo {getattr(file, 'name', 'XML')}: {e}")

    return registros


def generar_excel_xml(df):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Datos")

        worksheet = writer.sheets["Datos"]

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for col in worksheet.iter_cols(min_col=4, max_col=4, min_row=2):
            for cell in col:
                cell.number_format = '$#,##0.00'

        worksheet.column_dimensions["A"].width = 90
        worksheet.column_dimensions["B"].width = 15
        worksheet.column_dimensions["C"].width = 40
        worksheet.column_dimensions["D"].width = 18

    output.seek(0)
    return output


def mostrar_modulo_xml():
    st.title("Lector de XML CFDI a Excel")
    st.write("Sube uno o varios archivos XML para extraer CONCEPTO, FECHA, FOLIO e IMPORTE.")

    files = st.file_uploader(
        "Selecciona uno o varios archivos XML",
        type=["xml"],
        accept_multiple_files=True,
        key="xml_uploader"
    )

    if files:
        all_data = []

        for file in files:
            all_data.extend(parse_xml(file))

        if all_data:
            df = pd.DataFrame(all_data)

            st.subheader("Vista previa")
            df_vista = df.copy()
            df_vista["IMPORTE"] = df_vista["IMPORTE"].apply(lambda x: f"${x:,.2f}")
            st.dataframe(df_vista, use_container_width=True)

            output = generar_excel_xml(df)

            st.download_button(
                label="Descargar Excel",
                data=output,
                file_name="resultado_xml.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("No se encontraron datos válidos en los archivos XML.")
    else:
        st.info("Sube uno o varios archivos XML para comenzar.")


# =========================================================
# MÓDULO 2: TABLAS DE RENDIMIENTO
# =========================================================
def generar_desempeno():
    # Más 5, luego 4, pocos 3, nunca 2 ni 1
    return random.choices(
        population=[5, 4, 3],
        weights=[60, 35, 5],
        k=1
    )[0]


def primer_no_vacio_en_fila(ws, row_num):
    for cell in ws[row_num]:
        if normalizar_texto(cell.value):
            return normalizar_texto(cell.value)
    return ""


def color_hex_desde_fill(cell):
    fill = cell.fill
    if fill is None:
        return "D9D9D9"

    fg = fill.fgColor
    if fg is None:
        return "D9D9D9"

    rgb = getattr(fg, "rgb", None)
    if rgb:
        rgb = str(rgb)
        if len(rgb) == 8:
            return rgb[2:]
        if len(rgb) == 6:
            return rgb

    return "D9D9D9"


def es_participa(cell):
    """
    En tu archivo:
    - A = participa
    - R = no participa
    """
    valor = cell.value
    s = normalizar_texto(valor).strip().lower()
    return s == "a"


def leer_tabla_actividades(file_bytes):
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb.active

    titulo = primer_no_vacio_en_fila(ws, 1) or "TABLA DE ACTIVIDADES"
    puesto = primer_no_vacio_en_fila(ws, 2) or "PUESTO SIN DEFINIR"

    header_row = 3
    start_data_row = 5
    col_num = 1
    col_act = 2
    start_emp_col = 3

    empleados = []
    col = start_emp_col

    while col <= ws.max_column:
        nombre = normalizar_texto(ws.cell(header_row, col).value)
        if nombre == "":
            break

        color = color_hex_desde_fill(ws.cell(header_row, col))

        empleados.append({
            "nombre": nombre,
            "col": col,
            "color": color
        })
        col += 1

    actividades = []
    row = start_data_row

    while row <= ws.max_row:
        num = ws.cell(row, col_num).value
        actividad = normalizar_texto(ws.cell(row, col_act).value)

        if actividad == "":
            break

        participantes = []
        for emp in empleados:
            cell_part = ws.cell(row, emp["col"])
            if es_participa(cell_part):
                participantes.append({
                    "nombre": emp["nombre"],
                    "color": emp["color"]
                })

        actividades.append({
            "numero": num,
            "actividad": actividad,
            "participantes": participantes
        })

        row += 1

    return {
        "titulo": titulo,
        "puesto": puesto,
        "empleados": empleados,
        "actividades": actividades
    }


def generar_rendimiento(actividades, semanas=4):
    resultado = []

    for act in actividades:
        semanas_data = []
        for semana in range(1, semanas + 1):
            bloques = []
            for participante in act["participantes"]:
                bloques.append({
                    "nombre": participante["nombre"],
                    "color": participante["color"],
                    "valor": generar_desempeno()
                })

            semanas_data.append({
                "semana": semana,
                "bloques": bloques
            })

        resultado.append({
            "numero": act["numero"],
            "actividad": act["actividad"],
            "semanas": semanas_data
        })

    return resultado


def render_preview_html(puesto, mes_anio, rendimiento):
    html = f"""
    <div style="font-family:Arial,sans-serif; overflow-x:auto; margin-bottom:30px;">
      <table style="border-collapse:collapse; min-width:1100px; width:100%;">
        <tr>
          <th colspan="5" style="border:1px solid #000; background:#f2be00; font-size:22px; padding:8px;">
            TABLA DE RENDIMIENTO
          </th>
        </tr>
        <tr>
          <th colspan="5" style="border:1px solid #000; background:#f2be00; font-size:20px; padding:8px;">
            {puesto}
          </th>
        </tr>
        <tr>
          <th colspan="5" style="border:1px solid #000; background:#f2be00; font-size:20px; padding:8px;">
            {mes_anio}
          </th>
        </tr>
        <tr>
          <th style="border:1px solid #000; background:#eeeeee; width:100px; font-size:18px; padding:6px;">
            #ACT.
          </th>
          <th style="border:1px solid #000; background:#eeeeee; width:240px; font-size:18px;">
            SEMANA 1
          </th>
          <th style="border:1px solid #000; background:#eeeeee; width:240px; font-size:18px;">
            SEMANA 2
          </th>
          <th style="border:1px solid #000; background:#eeeeee; width:240px; font-size:18px;">
            SEMANA 3
          </th>
          <th style="border:1px solid #000; background:#eeeeee; width:240px; font-size:18px;">
            SEMANA 4
          </th>
        </tr>
    """

    for item in rendimiento:
        html += f"""
        <tr>
          <td style="border:1px solid #000; background:#f7f7f7; text-align:center; font-size:18px; padding:6px;">
            {item['numero']}
          </td>
        """

        for semana in item["semanas"]:
            bloques = semana["bloques"]

            if not bloques:
                html += '<td style="border:1px solid #000; height:38px; background:#fff;"></td>'
                continue

            html += '<td style="border:1px solid #000; padding:0; height:38px;">'
            html += '<div style="display:flex; width:100%; height:38px;">'

            for i, bloque in enumerate(bloques):
                borde = "border-right:1px solid #000;" if i < len(bloques) - 1 else ""
                html += f"""
                <div style="
                    flex:1;
                    background:#{bloque['color']};
                    {borde}
                    display:flex;
                    justify-content:center;
                    align-items:center;
                    font-size:18px;
                ">
                    {bloque['valor']}
                </div>
                """

            html += '</div></td>'

        html += "</tr>"

    html += "</table></div>"
    return html


def escribir_hoja_rendimiento(wb, nombre_hoja, puesto, mes_anio, rendimiento):
    limpiar_hoja_si_existe(wb, nombre_hoja)
    ws = wb.create_sheet(title=safe_sheet_name(nombre_hoja))

    yellow_fill = PatternFill(fill_type="solid", fgColor="F2BE00")
    gray_fill = PatternFill(fill_type="solid", fgColor="EFEFEF")
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")

    bloques_semana = {
        1: (2, 7),
        2: (8, 13),
        3: (14, 19),
        4: (20, 25)
    }

    ws.merge_cells("A1:Y1")
    ws["A1"] = "TABLA DE RENDIMIENTO"
    ws["A1"].fill = yellow_fill
    centrar(ws["A1"], bold=True, size=16)

    ws.merge_cells("A2:Y2")
    ws["A2"] = puesto
    ws["A2"].fill = yellow_fill
    centrar(ws["A2"], bold=True, size=15)

    ws.merge_cells("A3:Y3")
    ws["A3"] = mes_anio
    ws["A3"].fill = yellow_fill
    centrar(ws["A3"], bold=True, size=15)

    ws.merge_cells("A4:A5")
    ws["A4"] = "#ACT."
    ws["A4"].fill = gray_fill
    centrar(ws["A4"], bold=True, size=14)

    for semana, (c1, c2) in bloques_semana.items():
        ws.merge_cells(start_row=4, start_column=c1, end_row=4, end_column=c2)
        cell = ws.cell(4, c1)
        cell.value = f"SEMANA {semana}"
        cell.fill = gray_fill
        centrar(cell, bold=True, size=14)

        for col in range(c1, c2 + 1):
            ws.cell(5, col).fill = white_fill

    aplicar_borde_rango(ws, 1, 1, 5, 25)

    ajustar_ancho(ws, 1, 10)
    for col in range(2, 26):
        ajustar_ancho(ws, col, 5)

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 26
    ws.row_dimensions[5].height = 12

    start_row = 6

    for idx, item in enumerate(rendimiento, start=0):
        excel_row = start_row + idx
        ws.row_dimensions[excel_row].height = 28

        c = ws.cell(excel_row, 1)
        c.value = item["numero"]
        c.fill = white_fill
        centrar(c, size=14)
        c.border = Border(left=THIN_BLACK, right=THIN_BLACK, top=THIN_BLACK, bottom=THIN_BLACK)

        for sem_idx, semana in enumerate(item["semanas"], start=1):
            c1, c2 = bloques_semana[sem_idx]
            bloques = semana["bloques"]

            if len(bloques) == 0:
                ws.merge_cells(start_row=excel_row, start_column=c1, end_row=excel_row, end_column=c2)
                mc = ws.cell(excel_row, c1)
                mc.value = ""
                mc.fill = white_fill
                centrar(mc, size=14)
                aplicar_borde_rango(ws, excel_row, c1, excel_row, c2)

            elif len(bloques) == 1:
                b = bloques[0]
                ws.merge_cells(start_row=excel_row, start_column=c1, end_row=excel_row, end_column=c2)
                mc = ws.cell(excel_row, c1)
                mc.value = b["valor"]
                mc.fill = PatternFill(fill_type="solid", fgColor=b["color"])
                centrar(mc, size=14)
                aplicar_borde_rango(ws, excel_row, c1, excel_row, c2)

            elif len(bloques) == 2:
                rangos = [(c1, c1 + 2), (c1 + 3, c2)]
                for b, (r1, r2) in zip(bloques, rangos):
                    ws.merge_cells(start_row=excel_row, start_column=r1, end_row=excel_row, end_column=r2)
                    mc = ws.cell(excel_row, r1)
                    mc.value = b["valor"]
                    mc.fill = PatternFill(fill_type="solid", fgColor=b["color"])
                    centrar(mc, size=14)
                    aplicar_borde_rango(ws, excel_row, r1, excel_row, r2)

            else:
                total_cols = c2 - c1 + 1
                n = len(bloques)
                ancho_base = total_cols // n
                resto = total_cols % n

                rangos = []
                actual = c1

                for i in range(n):
                    ancho = ancho_base + (1 if i < resto else 0)
                    fin = actual + ancho - 1
                    rangos.append((actual, fin))
                    actual = fin + 1

                for b, (r1, r2) in zip(bloques, rangos):
                    ws.merge_cells(start_row=excel_row, start_column=r1, end_row=excel_row, end_column=r2)
                    mc = ws.cell(excel_row, r1)
                    mc.value = b["valor"]
                    mc.fill = PatternFill(fill_type="solid", fgColor=b["color"])
                    centrar(mc, size=14)
                    aplicar_borde_rango(ws, excel_row, r1, excel_row, r2)

    return wb


def construir_excel_final(file_bytes, puesto, mes1, anio1, mes2, anio2, rendimiento1, rendimiento2):
    wb = load_workbook(BytesIO(file_bytes))

    hoja1 = f"RENDIMIENTO {mes1} {anio1}"
    hoja2 = f"RENDIMIENTO {mes2} {anio2}"

    escribir_hoja_rendimiento(
        wb=wb,
        nombre_hoja=hoja1,
        puesto=puesto,
        mes_anio=f"{mes1} {anio1}",
        rendimiento=rendimiento1
    )

    escribir_hoja_rendimiento(
        wb=wb,
        nombre_hoja=hoja2,
        puesto=puesto,
        mes_anio=f"{mes2} {anio2}",
        rendimiento=rendimiento2
    )

    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida.getvalue()


def mostrar_modulo_rendimiento():
    st.title("Generador de dos tablas de rendimiento desde Excel")

    uploaded_file = st.file_uploader(
        "Sube el Excel con la TABLA DE ACTIVIDADES",
        type=["xlsx"],
        key="rendimiento_uploader"
    )

    meses = [
        "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
        "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
    ]

    st.subheader("Mes 1")
    c1, c2 = st.columns(2)
    with c1:
        mes1 = st.selectbox("Mes 1", meses, index=4, key="mes1")
    with c2:
        anio1 = st.number_input("Año 1", min_value=2024, max_value=2100, value=2025, step=1, key="anio1")

    st.subheader("Mes 2")
    c3, c4 = st.columns(2)
    with c3:
        mes2 = st.selectbox("Mes 2", meses, index=5, key="mes2")
    with c4:
        anio2 = st.number_input("Año 2", min_value=2024, max_value=2100, value=2025, step=1, key="anio2")

    if uploaded_file is not None:
        file_bytes = uploaded_file.read()

        try:
            data = leer_tabla_actividades(file_bytes)

            puesto = data["puesto"]
            empleados = data["empleados"]
            actividades = data["actividades"]

            st.success("Archivo leído correctamente.")
            st.write(f"**Puesto detectado:** {puesto}")
            st.write(f"**Empleados detectados:** {len(empleados)}")
            st.write(f"**Actividades detectadas:** {len(actividades)}")

            with st.expander("Ver empleados y colores detectados"):
                for emp in empleados:
                    st.markdown(
                        f"""
                        <div style="display:flex; align-items:center; gap:10px; margin-bottom:6px;">
                            <div style="width:22px; height:22px; background:#{emp['color']}; border:1px solid #000;"></div>
                            <div>{emp['nombre']}</div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )

            with st.expander("Ver participación detectada por actividad"):
                for act in actividades:
                    nombres = [p["nombre"] for p in act["participantes"]]
                    st.write(f"{act['numero']}. {act['actividad']} → {', '.join(nombres) if nombres else 'Sin participantes'}")

            col_btn1, col_btn2 = st.columns(2)

            with col_btn1:
                if st.button("Generar dos meses", key="btn_generar"):
                    st.session_state["rendimiento_mes_1"] = generar_rendimiento(actividades, semanas=4)
                    st.session_state["rendimiento_mes_2"] = generar_rendimiento(actividades, semanas=4)
                    st.session_state["file_bytes_original"] = file_bytes
                    st.session_state["puesto_detectado"] = puesto

            with col_btn2:
                if st.button("Regenerar aleatoriamente", key="btn_regenerar"):
                    st.session_state["rendimiento_mes_1"] = generar_rendimiento(actividades, semanas=4)
                    st.session_state["rendimiento_mes_2"] = generar_rendimiento(actividades, semanas=4)

            if "rendimiento_mes_1" in st.session_state and "rendimiento_mes_2" in st.session_state:
                st.subheader("Vista previa: Mes 1")
                html1 = render_preview_html(
                    puesto=st.session_state["puesto_detectado"],
                    mes_anio=f"{mes1} {anio1}",
                    rendimiento=st.session_state["rendimiento_mes_1"]
                )
                st.components.v1.html(html1, height=700, scrolling=True)

                st.subheader("Vista previa: Mes 2")
                html2 = render_preview_html(
                    puesto=st.session_state["puesto_detectado"],
                    mes_anio=f"{mes2} {anio2}",
                    rendimiento=st.session_state["rendimiento_mes_2"]
                )
                st.components.v1.html(html2, height=700, scrolling=True)

                excel_final = construir_excel_final(
                    file_bytes=st.session_state["file_bytes_original"],
                    puesto=st.session_state["puesto_detectado"],
                    mes1=mes1,
                    anio1=anio1,
                    mes2=mes2,
                    anio2=anio2,
                    rendimiento1=st.session_state["rendimiento_mes_1"],
                    rendimiento2=st.session_state["rendimiento_mes_2"]
                )

                st.download_button(
                    label="Descargar Excel final con dos tablas",
                    data=excel_final,
                    file_name=f"rendimiento_{mes1.lower()}_{anio1}_{mes2.lower()}_{anio2}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            st.error(f"Ocurrió un error al procesar el archivo: {e}")
    else:
        st.info("Sube el Excel de actividades para comenzar.")


# =========================================================
# MÓDULO 3: BITÁCORA DE ASISTENCIA
# =========================================================

# Texto -> color
STATUS_COLORS = {
    "LABORADO": "19A7D8",
    "SI LABORO": "19A7D8",
    "SI LABORÓ": "19A7D8",
    "NO LABORO": "C0504D",
    "NO LABORÓ": "C0504D",
    "INHABIL": "FF0000",
    "INHÁBIL": "FF0000",
    "INCAPACIDAD": "59A8C0",
    "VACACIONES": "8C8C8C",
    "BAJA": "FFFF00",
    "NO HABIA INGRESADO": "FFFFFF",
    "NO HABÍA INGRESADO": "FFFFFF"
}

STATUS_LABELS = [
    ("SI LABORÓ", "19A7D8"),
    ("NO LABORÓ", "C0504D"),
    ("INHÁBIL", "FF0000"),
    ("INCAPACIDAD", "59A8C0"),
    ("VACACIONES", "8C8C8C"),
    ("BAJA", "FFFF00"),
    ("NO HABÍA INGRESADO", "FFFFFF"),
]

DAY_LETTERS = {
    0: "L",
    1: "M",
    2: "M",
    3: "J",
    4: "V",
    5: "S",
    6: "D",
}

MONTH_NAMES_ES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}


def nth_weekday_of_month(year, month, weekday, n):
    """
    weekday: lunes=0 ... domingo=6
    """
    first_day = date(year, month, 1)
    days_until_weekday = (weekday - first_day.weekday()) % 7
    first_occurrence = first_day + timedelta(days=days_until_weekday)
    return first_occurrence + timedelta(weeks=n - 1)


def get_official_mexico_holidays(year):
    holidays = set()

    holidays.add(date(year, 1, 1))
    holidays.add(nth_weekday_of_month(year, 2, 0, 1))
    holidays.add(nth_weekday_of_month(year, 3, 0, 3))
    holidays.add(date(year, 5, 1))
    holidays.add(date(year, 9, 16))
    holidays.add(nth_weekday_of_month(year, 11, 0, 3))

    if year >= 2024 and (year - 2024) % 6 == 0:
        holidays.add(date(year, 10, 1))

    holidays.add(date(year, 12, 25))

    return holidays


def normalize_text(text):
    if text is None:
        return ""
    text = str(text).strip().upper()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = " ".join(text.split())
    return text


def parse_excel_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    if isinstance(value, str):
        value = value.strip()
        formats = [
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%Y-%m-%d",
            "%d-%m-%y",
            "%d/%m/%y",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass
    return None


def find_header_row_and_columns(ws):
    empleado_col = None
    cliente_col = None
    header_row = None
    date_cols = []

    for row in range(1, min(ws.max_row, 30) + 1):
        row_values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        normalized = [normalize_text(v) for v in row_values]

        if "EMPLEADO" in normalized:
            header_row = row
            empleado_col = normalized.index("EMPLEADO") + 1
            if "CLIENTE" in normalized:
                cliente_col = normalized.index("CLIENTE") + 1

            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=header_row, column=col).value
                if parse_excel_date(cell_value):
                    date_cols.append(col)

            break

    if not header_row or not empleado_col or not date_cols:
        raise ValueError(
            "No se pudo identificar la estructura del archivo. "
            "Verifica que exista una fila con 'Empleado' y encabezados de fechas."
        )

    return header_row, empleado_col, cliente_col, date_cols


def split_employee_code_name(raw_name):
    text = "" if raw_name is None else str(raw_name).strip()
    text = " ".join(text.split())
    parts = re.split(r"\s*-\s*", text, maxsplit=1)

    if len(parts) == 2:
        code = parts[0].strip()
        name = parts[1].strip()
        return code, name

    return "", text


def build_employee_index(ws, header_row, empleado_col, cliente_col, date_cols):
    employees = {}

    for row in range(header_row + 1, ws.max_row + 1):
        raw_name = ws.cell(row=row, column=empleado_col).value
        if raw_name is None:
            continue

        original_name = " ".join(str(raw_name).strip().split())
        if not original_name:
            continue

        employee_code, clean_name = split_employee_code_name(original_name)

        norm_original = normalize_text(original_name)
        norm_clean = normalize_text(clean_name)
        norm_code = normalize_text(employee_code)

        cliente = ws.cell(row=row, column=cliente_col).value if cliente_col else ""

        statuses = {}
        for col in date_cols:
            fecha = parse_excel_date(ws.cell(row=header_row, column=col).value)
            estado = ws.cell(row=row, column=col).value
            if fecha:
                statuses[fecha] = normalize_text(estado)

        data = {
            "row": row,
            "original_name": original_name,
            "clean_name": clean_name,
            "employee_code": employee_code,
            "cliente": str(cliente).strip() if cliente else "",
            "statuses": statuses
        }

        if norm_original:
            employees[norm_original] = data
        if norm_clean:
            employees[norm_clean] = data
        if norm_code:
            employees[norm_code] = data

    return employees


def find_employee(employee_index, search_name):
    search = normalize_text(search_name)

    if not search:
        return None

    if search in employee_index:
        return employee_index[search]

    partial_matches = []
    seen_rows = set()

    for _, data in employee_index.items():
        row_id = data["row"]
        if row_id in seen_rows:
            continue

        seen_rows.add(row_id)

        original_norm = normalize_text(data.get("original_name", ""))
        clean_norm = normalize_text(data.get("clean_name", ""))
        code_norm = normalize_text(data.get("employee_code", ""))

        if (
            search in clean_norm
            or clean_norm in search
            or search in original_norm
            or search in code_norm
        ):
            partial_matches.append(data)

    if len(partial_matches) == 1:
        return partial_matches[0]

    return None


def get_months_present(date_list):
    return sorted(set((d.year, d.month) for d in date_list))


def get_pre_month_default_status(statuses, year, month):
    month_days = sorted(
        d for d in statuses.keys()
        if d.year == year and d.month == month and d.weekday() <= 4
    )

    for d in month_days:
        status = normalize_text(statuses.get(d, ""))
        if status:
            if status == "NO HABIA INGRESADO":
                return "NO HABÍA INGRESADO"
            return "SI LABORÓ"

    return "SI LABORÓ"


def status_to_fill(status):
    color = STATUS_COLORS.get(normalize_text(status), "FFFFFF")
    return PatternFill(fill_type="solid", fgColor=color)


def apply_border(cell):
    thin = Side(style="thin", color="000000")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def set_cell(ws, row, col, value=None, fill=None, font=None, alignment=None, merge=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    apply_border(cell)

    if merge:
        ws.merge_cells(start_row=row, start_column=col, end_row=merge[0], end_column=merge[1])

    return cell


def get_company_name(ws_in):
    company_name = "EMPRESA"
    for row in range(1, min(ws_in.max_row, 15) + 1):
        for col in range(1, min(ws_in.max_column, 8) + 1):
            val = ws_in.cell(row=row, column=col).value
            if val and "EMPRESA:" in normalize_text(val):
                company_name = str(val).replace("Empresa:", "").replace("EMPRESA:", "").strip()
                return company_name
    return company_name


def get_selected_employees(employee_index, employee_names):
    selected = []
    seen_rows = set()
    not_found = []

    for emp_name in employee_names:
        emp_name = emp_name.strip()
        if not emp_name:
            continue

        data = find_employee(employee_index, emp_name)
        if not data:
            not_found.append(emp_name)
            continue

        row_id = data["row"]
        if row_id not in seen_rows:
            selected.append(data)
            seen_rows.add(row_id)

    return selected, not_found


def group_days_into_weeks(year, month):
    first_day = date(year, month, 1)

    if first_day.weekday() <= 4:
        start_monday = first_day - timedelta(days=first_day.weekday())
    else:
        start_monday = first_day + timedelta(days=(7 - first_day.weekday()))

    weeks = []
    current_monday = start_monday

    for _ in range(4):
        week = []
        for i in range(5):
            week.append(current_monday + timedelta(days=i))
        weeks.append(week)
        current_monday += timedelta(days=7)

    return weeks


def set_range_border(ws, start_row, start_col, end_row, end_col):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            apply_border(ws.cell(r, c))


def create_month_table(ws_out, start_row, company_name, year, month, selected_employees):
    red_fill = PatternFill(fill_type="solid", fgColor="FF0000")
    light_header_fill = PatternFill(fill_type="solid", fgColor="EAD8C8")
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    bold_white = Font(bold=True, color="FFFFFF", size=12)
    bold_black = Font(bold=True, color="000000", size=11)
    title_font = Font(bold=True, color="FF0000", size=18)
    normal_font = Font(size=11, color="000000")

    widths = {
        1: 4,
        2: 16,
        3: 28,
    }
    for c in range(4, 40):
        widths[c] = 5

    for col, width in widths.items():
        ws_out.column_dimensions[get_column_letter(col)].width = width

    weeks = group_days_into_weeks(year, month)

    start_col = 4
    col = start_col
    week_ranges = []

    for week in weeks[:4]:
        end_col = col + 5 - 1
        week_ranges.append((col, end_col, week))
        col = end_col + 1

    last_col = max(14, week_ranges[-1][1])

    if start_row == 2:
        ws_out.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=last_col)
        ws_out.cell(start_row, 2, "BITÁCORA DE ASISTENCIA")
        ws_out.cell(start_row, 2).font = title_font
        ws_out.cell(start_row, 2).alignment = center

    ws_out.merge_cells(start_row=start_row + 2, start_column=2, end_row=start_row + 2, end_column=last_col)
    set_cell(ws_out, start_row + 2, 2, company_name, fill=red_fill, font=bold_white, alignment=center)

    month_name = f"{MONTH_NAMES_ES[month]} {year}"
    ws_out.merge_cells(start_row=start_row + 3, start_column=2, end_row=start_row + 3, end_column=last_col)
    set_cell(ws_out, start_row + 3, 2, month_name, fill=white_fill, font=Font(bold=True, size=14), alignment=center)

    ws_out.merge_cells(start_row=start_row + 4, start_column=2, end_row=start_row + 4, end_column=last_col)
    set_cell(ws_out, start_row + 4, 2, "LISTA DE ASISTENCIA", fill=red_fill, font=bold_white, alignment=center)

    header_row_1 = start_row + 5
    header_row_2 = start_row + 6

    ws_out.merge_cells(start_row=header_row_1, start_column=2, end_row=header_row_2, end_column=2)
    ws_out.merge_cells(start_row=header_row_1, start_column=3, end_row=header_row_2, end_column=3)

    set_cell(ws_out, header_row_1, 2, "NUM DE\nEMPLEADO", fill=white_fill, font=Font(bold=True, size=11), alignment=center)
    set_cell(ws_out, header_row_1, 3, "NOMBRE\nCOMPLETO", fill=white_fill, font=Font(bold=True, size=11), alignment=center)

    for i, (week_start, week_end, week_days) in enumerate(week_ranges, start=1):
        ws_out.merge_cells(start_row=header_row_1, start_column=week_start, end_row=header_row_1, end_column=week_end)
        set_cell(ws_out, header_row_1, week_start, f"SEMANA {i}", fill=light_header_fill, font=bold_black, alignment=center)
        set_range_border(ws_out, header_row_1, week_start, header_row_1, week_end)

        for offset, day in enumerate(week_days):
            cell = ws_out.cell(row=header_row_2, column=week_start + offset, value=DAY_LETTERS[day.weekday()])
            cell.font = bold_black
            cell.alignment = center
            apply_border(cell)

    data_start_row = start_row + 7
    current_row = data_start_row

    for index, emp in enumerate(selected_employees, start=1):
        employee_number = str(index)
        employee_name = emp.get("clean_name", emp.get("original_name", ""))

        set_cell(ws_out, current_row, 2, employee_number, fill=white_fill, font=Font(bold=True), alignment=center)
        set_cell(ws_out, current_row, 3, employee_name, fill=white_fill, font=normal_font, alignment=left_align)

        first_day_of_month = date(year, month, 1)
        official_holidays = get_official_mexico_holidays(year)
        pre_month_status = get_pre_month_default_status(emp["statuses"], year, month)

        for week_start, week_end, week_days in week_ranges:
            for offset, day in enumerate(week_days):
                if day < first_day_of_month:
                    fill = status_to_fill(pre_month_status)

                elif day.year == year and day.month == month:
                    status = normalize_text(emp["statuses"].get(day, ""))

                    if day in official_holidays:
                        fill = status_to_fill("INHÁBIL")
                    elif status == "NO HABIA INGRESADO":
                        fill = status_to_fill("NO HABÍA INGRESADO")
                    else:
                        fill = status_to_fill(status)

                else:
                    fill = white_fill

                set_cell(ws_out, current_row, week_start + offset, "", fill=fill, alignment=center)

        current_row += 1

    legend_row = current_row + 2
    legend_col = 2

    for label, color in STATUS_LABELS:
        ws_out.merge_cells(
            start_row=legend_row,
            start_column=legend_col,
            end_row=legend_row + 1,
            end_column=legend_col + 1
        )
        c = ws_out.cell(row=legend_row, column=legend_col, value=label)
        c.fill = PatternFill(fill_type="solid", fgColor=color)
        c.alignment = center
        c.font = Font(bold=False, color="000000")
        apply_border(c)

        for r in range(legend_row, legend_row + 2):
            for cc in range(legend_col, legend_col + 2):
                apply_border(ws_out.cell(r, cc))

        legend_col += 3

    ws_out.row_dimensions[start_row].height = 28
    ws_out.row_dimensions[start_row + 2].height = 24
    ws_out.row_dimensions[start_row + 3].height = 24
    ws_out.row_dimensions[start_row + 4].height = 24
    ws_out.row_dimensions[header_row_1].height = 42
    ws_out.row_dimensions[header_row_2].height = 24

    for r in range(data_start_row, current_row):
        ws_out.row_dimensions[r].height = 38

    return legend_row + 4


def generate_report(input_file, employee_names, output_file):
    wb_in = load_workbook(input_file, data_only=True)
    ws_in = wb_in.active

    header_row, empleado_col, cliente_col, date_cols = find_header_row_and_columns(ws_in)
    employee_index = build_employee_index(ws_in, header_row, empleado_col, cliente_col, date_cols)

    company_name = get_company_name(ws_in)
    selected_employees, not_found = get_selected_employees(employee_index, employee_names)

    if not selected_employees:
        raise ValueError("No se encontró ningún empleado con los nombres capturados.")

    all_dates = sorted(selected_employees[0]["statuses"].keys())
    months_present = get_months_present(all_dates)

    if not months_present:
        raise ValueError("No se encontraron fechas válidas en el archivo.")

    bimester_months = months_present[:2]

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Bitácora"

    next_row = 2
    first_table = True

    for year, month in bimester_months:
        if not first_table:
            next_row += 2

        next_row = create_month_table(
            ws_out=ws_out,
            start_row=next_row,
            company_name=company_name,
            year=year,
            month=month,
            selected_employees=selected_employees
        )

        first_table = False

    ws_out.freeze_panes = "D8"
    wb_out.save(output_file)
    return not_found


def mostrar_modulo_bitacora():
    st.title("Generador de Bitácora de Asistencia")

    st.write(
        "Escribe uno o varios nombres exactamente como aparecen en el Excel. "
        "También puede encontrar coincidencias parciales si solo coincide un empleado."
    )

    empleados_texto = st.text_area(
        "Nombres de empleados",
        height=180,
        placeholder="Empleado 1\nEmpleado 2\nEmpleado 3",
        key="bitacora_empleados"
    )

    archivo_excel = st.file_uploader(
        "Selecciona el archivo de asistencias",
        type=["xlsx", "xlsm", "xltx", "xltm"],
        key="bitacora_uploader"
    )

    nombre_salida = st.text_input(
        "Nombre del archivo de salida",
        value="bitacora_asistencia.xlsx",
        key="bitacora_salida"
    )

    if st.button("Generar Excel", key="bitacora_generar"):
        if not archivo_excel:
            st.error("Selecciona el archivo de asistencias.")
            return

        employee_names = [x.strip() for x in empleados_texto.splitlines() if x.strip()]
        if not employee_names:
            st.error("Escribe al menos un nombre de empleado.")
            return

        if not nombre_salida.strip():
            nombre_salida = "bitacora_asistencia.xlsx"

        if not nombre_salida.lower().endswith(".xlsx"):
            nombre_salida += ".xlsx"

        input_path = None
        output_path = None

        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_in:
                temp_in.write(archivo_excel.getbuffer())
                input_path = temp_in.name

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_out:
                output_path = temp_out.name

            not_found = generate_report(input_path, employee_names, output_path)

            with open(output_path, "rb") as f:
                output_bytes = f.read()

            st.success("Archivo generado correctamente.")

            if not_found:
                st.warning(
                    "No se encontraron estos empleados:\n- " + "\n- ".join(not_found)
                )

            st.download_button(
                label="Descargar reporte Excel",
                data=output_bytes,
                file_name=nombre_salida,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Ocurrió un problema:\n\n{str(e)}")

        finally:
            try:
                if input_path and os.path.exists(input_path):
                    os.remove(input_path)
            except Exception:
                pass

            try:
                if output_path and os.path.exists(output_path):
                    os.remove(output_path)
            except Exception:
                pass


# =========================================================
# SIDEBAR / NAVEGACIÓN
# =========================================================
st.sidebar.title("Menú")
opcion = st.sidebar.radio(
    "¿Qué quieres hacer?",
    [
        "XML CFDI a Excel",
        "Tablas de rendimiento",
        "Bitácora de asistencia"
    ]
)

st.sidebar.markdown("---")
st.sidebar.write("Selecciona una opción para trabajar.")


# =========================================================
# ROUTER PRINCIPAL
# =========================================================
if opcion == "XML CFDI a Excel":
    mostrar_modulo_xml()
elif opcion == "Tablas de rendimiento":
    mostrar_modulo_rendimiento()
elif opcion == "Bitácora de asistencia":
    mostrar_modulo_bitacora()